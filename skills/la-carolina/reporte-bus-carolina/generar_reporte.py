#!/usr/bin/env python3
"""
Genera un reporte (1 página, branding La Carolina) de un BUS o un CONDUCTOR.
Uso:
    python3 generar_reporte.py bus 622
    python3 generar_reporte.py conductor "HURTADO QUIROGA"
    python3 generar_reporte.py conductor 72282385      (por cédula)

Lee los archivos operativos de La Carolina (local OneDrive). Genera HTML + PDF.
"""
import sys, os, csv, subprocess, re, shutil
from collections import Counter, defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl: pip3 install openpyxl")

# ---------- RUTAS ----------
# Adaptado para Hermes/Linux. Define LC_MTC_DIR si los Excel están en una ruta distinta.
DEFAULT_MTC_CANDIDATES = [
    os.environ.get("LC_MTC_DIR", ""),
    "/root/OneDrive/Transporte/MTC",
    "/root/obsidian-vault/empresas/la-carolina/Operacion y flota/AFILIADOS 2026",
]
MTC = next((p for p in DEFAULT_MTC_CANDIDATES if p and os.path.isdir(p)), os.environ.get("LC_MTC_DIR", DEFAULT_MTC_CANDIDATES[1]))
_viajes_xlsx = os.path.join(MTC, "00_INBOX MTC", "Viajes perdidos.xlsx")
_viajes_csv = os.path.join(MTC, "Consolidados_Operativo", "Operativo_Vjperdidos_consolidado.csv")
_vehcond_xlsx = os.path.join(MTC, "00_INBOX MTC", "VEHICULO - CONDUCTORES  A MAYO 2026.xlsx")
_vehcond_csv = os.path.join(MTC, "Consolidados_Operativo", "Base_comision_Conductores_consolidado.csv")
F_VIAJES = os.environ.get("LC_VIAJES_XLSX") or (_viajes_xlsx if os.path.exists(_viajes_xlsx) else _viajes_csv)
F_VEHCOND = os.environ.get("LC_VEHCOND_XLSX") or (_vehcond_xlsx if os.path.exists(_vehcond_xlsx) else _vehcond_csv)
SKILL_DIR = os.path.dirname(os.path.abspath(__file__))
F_MODELO = os.path.join(SKILL_DIR, "data", "flota_modelo.csv")
OUT_DIR = os.environ.get("LC_AFILIADOS_REPORT_OUT") or os.path.join(MTC, "AFILIADOS 2026", "Reportes por buseta y conductor")
CHROME = (os.environ.get("LC_CHROME") or shutil.which("chromium-browser") or shutil.which("chromium") or
          shutil.which("google-chrome") or shutil.which("google-chrome-stable") or
          "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")

def require_file(path, label):
    if not os.path.exists(path):
        sys.exit(f"No encuentro {label}: {path}\n"
                 "Define LC_MTC_DIR al directorio Transporte/MTC o LC_VIAJES_XLSX/LC_VEHCOND_XLSX con rutas exactas.")

def require_runtime_inputs():
    require_file(F_VIAJES, "Viajes perdidos.xlsx")
    require_file(F_VEHCOND, "VEHICULO - CONDUCTORES A MAYO 2026.xlsx")
    require_file(F_MODELO, "data/flota_modelo.csv")
    if not CHROME or not os.path.exists(CHROME):
        sys.exit("No encuentro Chrome/Chromium para generar PDF. Instala chromium-browser o define LC_CHROME.")

DIAS = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo','Sin fecha']
MESES = {'01':'Enero','02':'Febrero','03':'Marzo','04':'Abril','05':'Mayo','06':'Junio'}
VEH_CAUSAS = {'TALLER','VARADO EN RUTA'}
TIMBRADAS_POR_VUELTA_FLOTA = 75  # promedio flota (fallback)
TARIFA_TIMBRADA = 3_300          # COP por timbrada (tarifa vigente)
FLOTA_VEH_PROM = 57              # vueltas perdidas por vehículo promedio flota
PERIODO_INICIO = datetime(2026, 1, 1)
PERIODO_FIN = datetime(2026, 5, 31, 23, 59, 59)

def parse_date(v):
    if isinstance(v, datetime): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d")
    except: return None

def en_periodo(d):
    return bool(d and PERIODO_INICIO <= d <= PERIODO_FIN)

def norm_bus(v):
    return str(v).strip().replace('.0', '') if v is not None else ''

def norm_num(v):
    try: return float(str(v).replace(',', '').strip())
    except: return 0.0

def iter_vehcond():
    """Filas normalizadas de liquidación/vehículo-conductor. Soporta XLSX legado y CSV consolidado de Drive."""
    if F_VEHCOND.lower().endswith('.csv'):
        with open(F_VEHCOND, newline='', encoding='utf-8-sig') as f:
            for r in csv.DictReader(f, delimiter=';'):
                d = parse_date(r.get('FECHA'))
                if not en_periodo(d):
                    continue
                yield dict(
                    bus=norm_bus(r.get('VEHICULO')),
                    prop='',
                    cond=str(r.get('CONDUCTOR') or '').strip(),
                    cc=str(r.get('COD CONDUCTOR') or '').strip(),
                    fecha=d,
                    ruta=str(r.get('RUTA') or '').strip(),
                    viajes=int(norm_num(r.get('VIAJES'))),
                    timbradas=norm_num(r.get('TIMBRADAS')),
                )
        return
    wb = openpyxl.load_workbook(F_VEHCOND, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=6, values_only=True):
        d = parse_date(r[2] if len(r) > 2 else None)
        if not en_periodo(d):
            continue
        yield dict(
            bus=norm_bus(r[9] if len(r) > 9 else None),
            prop=str(r[11]).strip() if len(r) > 11 and r[11] else '',
            cond=str((r[8] if len(r) > 8 else None) or (r[7] if len(r) > 7 else '') or '').strip(),
            cc=str(r[7]).strip() if len(r) > 7 and r[7] else '',
            fecha=d,
            ruta=str(r[3]).strip() if len(r) > 3 and r[3] else '',
            viajes=int(norm_num(r[13] if len(r) > 13 else 0)),
            timbradas=norm_num(r[15] if len(r) > 15 else 0),
        )
    wb.close()

def iter_viajes():
    """Filas normalizadas de viajes perdidos. Soporta XLSX legado y CSV consolidado de Drive."""
    if F_VIAJES.lower().endswith('.csv'):
        with open(F_VIAJES, newline='', encoding='utf-8-sig') as f:
            for r in csv.DictReader(f, delimiter=';'):
                d = parse_date(r.get('Fecha'))
                if not en_periodo(d):
                    continue
                yield dict(
                    bus=norm_bus(r.get('Vehiculo')),
                    nov=str(r.get('Novedad') or '').strip(),
                    detalle=str(r.get('Detalle Novedad') or '').strip(),
                    fecha=d,
                    conductor=str(r.get('Conductor') or '').strip(),
                    cc=str(r.get('Ced. Conductor') or '').strip(),
                    viaje=r.get('Viaje'),
                    ruta=str(r.get('Ruta') or '').strip(),
                    placa=str(r.get('Placa') or '').strip(),
                )
        return
    wb = openpyxl.load_workbook(F_VIAJES, read_only=True, data_only=True)
    ws = wb['Base D']
    for r in ws.iter_rows(min_row=6, values_only=True):
        d = parse_date(r[4] if len(r) > 4 else None)
        if not en_periodo(d):
            continue
        yield dict(
            bus=norm_bus(r[7] if len(r) > 7 else None),
            nov=str(r[2]).strip() if len(r) > 2 and r[2] else '',
            detalle=str(r[3]).strip() if len(r) > 3 and r[3] else '',
            fecha=d,
            conductor=str(r[9]).strip() if len(r) > 9 and r[9] else '',
            cc=str(r[10]).strip() if len(r) > 10 and r[10] else '',
            viaje=r[12] if len(r) > 12 else None,
            ruta=str(r[13]).strip() if len(r) > 13 and r[13] else '',
            placa=str(r[8]).strip() if len(r) > 8 and r[8] else '',
        )
    wb.close()

def load_modelo():
    m, p = {}, {}
    if os.path.exists(F_MODELO):
        for row in csv.reader(open(F_MODELO)):
            if len(row) >= 3:
                m[row[0]] = row[2]; p[row[0]] = row[1]
    return m, p

# ---------- ANÁLISIS DE BUS ----------
def analizar_bus(bus):
    modelo, placa = load_modelo()
    bus = norm_bus(bus)
    prop=""; conductores=set(); dias=set(); rutas=Counter(); cond_dates=defaultdict(set)
    cond_mes=defaultdict(set)  # mes_key → set de conductores
    vj_sem=0; tim_sem=0.0; vj_fds=0; tim_fds=0.0

    for r in iter_vehcond():
        if r['bus'] != bus: continue
        if r.get('prop'): prop = r['prop']
        if r.get('cond'):
            conductores.add(r['cond'])
            cond_dates[r['cond']].add(r['fecha'].strftime('%Y-%m-%d'))
            cond_mes[r['fecha'].strftime('%Y-%m')].add(r['cond'])
        dias.add(r['fecha'].strftime('%Y-%m-%d'))
        if r.get('ruta'): rutas[r['ruta']] += 1
        vj = int(r.get('viajes') or 0); tm = float(r.get('timbradas') or 0)
        if vj > 0 and tm > 5:
            if r['fecha'].weekday() >= 5:
                vj_fds += vj; tim_fds += tm
            else:
                vj_sem += vj; tim_sem += tm

    ndias_op = len(dias)
    ruta = rutas.most_common(1)[0][0] if rutas else "—"
    conductor_fijo = max(cond_dates, key=lambda c: len(cond_dates[c])) if cond_dates else "—"
    conductor_fijo_dias = len(cond_dates[conductor_fijo]) if conductor_fijo != "—" else 0
    tim_sem_vuelta = round(tim_sem/vj_sem, 1) if vj_sem > 0 else TIMBRADAS_POR_VUELTA_FLOTA
    tim_fds_vuelta = round(tim_fds/vj_fds, 1) if vj_fds > 0 else TIMBRADAS_POR_VUELTA_FLOTA

    causas=Counter(); aus_det=Counter(); taller_dias=set()
    por_mes=Counter(); por_dia=Counter(); por_vuelta=Counter(); por_ruta=Counter(); fechas=[]; aus_fechas=[]
    timbradas_perdidas=0; placa_viajes=''
    for r in iter_viajes():
        if r['bus'] != bus: continue
        nov = r['nov']; causas[nov]+=1
        d = r['fecha']; fechas.append(d); por_mes[d.strftime("%m")]+=1; por_dia[d.weekday()]+=1
        if r.get('placa'): placa_viajes = r['placa']
        if nov=='TALLER': taller_dias.add(d.date())
        if nov in VEH_CAUSAS:
            rate = tim_fds_vuelta if d.weekday() >= 5 else tim_sem_vuelta
            timbradas_perdidas += rate
        try: por_vuelta[int(float(str(r.get('viaje'))))]+=1
        except: por_vuelta[0]+=1
        ruta_raw = re.sub(r'-+', '-', r.get('ruta') or 'Sin ruta')
        por_ruta[ruta_raw]+=1
        if nov=='AUSENCIA CONDUCTOR':
            aus_det[r.get('detalle') or '(sin detalle)']+=1
            aus_fechas.append(d)

    total_perd = sum(causas.values())
    veh_perd = sum(causas[c] for c in VEH_CAUSAS)
    aus_perd = causas.get('AUSENCIA CONDUCTOR',0)
    rango = (min(fechas).strftime("%d/%m/%Y"), max(fechas).strftime("%d/%m/%Y")) if fechas else ("—","—")
    aus_rango = (min(aus_fechas).strftime("%d/%m/%Y"), max(aus_fechas).strftime("%d/%m/%Y")) if aus_fechas else ("—","—")
    mod = modelo.get(bus,"—"); edad = (2026-int(mod)) if mod.isdigit() else "—"

    timbradas_perdidas = round(timbradas_perdidas)
    ingreso_perdido = timbradas_perdidas * TARIFA_TIMBRADA
    return dict(tipo='bus', bus=bus, prop=prop or "—", placa=placa.get(bus) or placa_viajes or "—",
        modelo=mod, edad=edad, ruta=ruta, ndias_op=ndias_op, ndias_sin=151-ndias_op,
        conductores=len(conductores), conductor_fijo=conductor_fijo, conductor_fijo_dias=conductor_fijo_dias,
        cond_mes=cond_mes,
        total_perd=total_perd, veh_perd=veh_perd, aus_perd=aus_perd,
        causas=causas, aus_det=aus_det, taller=len(taller_dias),
        por_mes=por_mes, por_dia=por_dia, por_vuelta=por_vuelta, por_ruta=por_ruta, rango=rango, aus_rango=aus_rango,
        tim_sem=tim_sem_vuelta, tim_fds=tim_fds_vuelta,
        timbradas_perdidas=timbradas_perdidas, ingreso_perdido=ingreso_perdido)

# ---------- ANÁLISIS DE CONDUCTOR ----------
def analizar_conductor(query):
    q = query.upper().strip()
    nombre=""; ced=""; buses=Counter(); dias=set(); rutas=Counter()
    for r in iter_vehcond():
        nom = r.get('cond') or ''; cc = r.get('cc') or ''
        if q in nom.upper() or (q.isdigit() and q==cc):
            nombre = nom; ced = cc
            if r.get('bus'): buses[r['bus']]+=1
            dias.add(r['fecha'].strftime('%Y-%m-%d'))
            if r.get('ruta'): rutas[r['ruta']]+=1
    if not nombre: return None

    causas=Counter(); aus_det=Counter(); fechas=[]
    for r in iter_viajes():
        nom = r.get('conductor') or ''; cc = r.get('cc') or ''
        if q in nom.upper() or (q.isdigit() and q==cc):
            nov=r['nov']; causas[nov]+=1
            d=r['fecha']; fechas.append(d)
            if nov=='AUSENCIA CONDUCTOR':
                aus_det[r.get('detalle') or '(sin detalle)']+=1
    rango = (min(fechas).strftime("%d/%m/%Y"), max(fechas).strftime("%d/%m/%Y")) if fechas else ("—","—")
    return dict(tipo='conductor', nombre=nombre, ced=ced, buses=buses, ndias=len(dias),
        rutas=rutas, causas=causas, aus_det=aus_det, total_perd=sum(causas.values()), rango=rango)

# ---------- HTML (branding La Carolina, 1 página) ----------
CSS = """@page{size:A4 portrait;margin:0}
:root{--dorado:#DCBE61;--rojo:#C22219;--negro:#222222;--gris:#F5F5F5;--borde:#E5E5E5}
*{box-sizing:border-box;margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact}
body{font-family:'Montserrat',sans-serif;color:#222;line-height:1.45}
.bebas{font-family:'Bebas Neue','Impact',sans-serif}
.page{width:210mm;height:297mm;margin:0 auto;display:flex;flex-direction:column;overflow:hidden}
.bar{height:6px;background:linear-gradient(90deg,#DCBE61 0 50%,#C22219 50% 100%)}
.hero{background:#222;color:#fff;padding:11px 22px;position:relative;overflow:hidden}
.hero::before{font-family:'Bebas Neue';position:absolute;font-size:170px;color:#DCBE61;opacity:.10;right:-10px;top:-45px;line-height:.7}
.hero-row{display:flex;justify-content:space-between;align-items:flex-start;position:relative;z-index:2}
.eyebrow{display:inline-block;background:#C22219;color:#fff;font-family:'Bebas Neue';font-size:11px;letter-spacing:2px;padding:3px 10px;margin-bottom:6px}
.hero h1{font-family:'Bebas Neue';font-size:34px;line-height:.95}.hero h1 .r{color:#DCBE61}
.hero .slogan{font-family:'Bebas Neue';color:#DCBE61;font-size:15px;border-left:3px solid #C22219;padding-left:9px;margin-top:5px}
.hero .meta{text-align:right;font-family:'Bebas Neue';font-size:11px;letter-spacing:1.5px;color:rgba(255,255,255,.65);line-height:1.5}
.body{padding:8px 22px 0}
.daterange{background:#F5F5F5;border-left:4px solid #DCBE61;padding:5px 12px;font-size:10.5px;color:#444;margin-bottom:7px}
.daterange b{color:#222}
h2{font-family:'Bebas Neue';font-size:18px;margin:9px 0 4px;border-bottom:3px solid #C22219;display:inline-block;padding-bottom:2px}
.ficha{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.f{background:#F5F5F5;border-radius:3px;padding:6px 12px;display:flex;justify-content:space-between;font-size:12px}
.f b{font-family:'Bebas Neue';letter-spacing:1px;color:#555;font-weight:400}.f .v{font-weight:700}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-top:8px}
.k{background:#F5F5F5;border-top:5px solid #DCBE61;border-radius:3px;padding:8px;text-align:center}
.k.r{border-top-color:#C22219}.k .v{font-family:'Bebas Neue';font-size:30px;line-height:.85}.k .l{font-size:9.5px;color:#555;margin-top:5px;line-height:1.25;font-weight:500}
table{width:100%;border-collapse:collapse;font-size:11px;margin-top:4px}
th{background:#222;color:#DCBE61;font-family:'Bebas Neue';letter-spacing:1px;padding:4px 9px;text-align:left;font-weight:400}
td{padding:3.5px 9px;border-bottom:1px solid #E5E5E5}.n{text-align:right;font-weight:700}
tr:nth-child(even) td{background:#F5F5F5}
.row{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:4px}
.box{border:1px solid #E5E5E5;border-radius:4px;padding:9px 13px}.box.r{border-top:5px solid #C22219}.box.d{border-top:5px solid #DCBE61}
.box h3{font-family:'Bebas Neue';font-size:15px;margin-bottom:5px}
.diag{background:#DCBE61;border-radius:4px;padding:9px 16px;margin-top:8px}.diag h3{font-family:'Bebas Neue';font-size:18px;margin-bottom:5px}
.diag ul{list-style:none;font-size:11px;line-height:1.45}.diag li{padding-left:16px;position:relative;margin:2px 0}
.diag li::before{content:"»";color:#C22219;position:absolute;left:0;font-family:'Bebas Neue';font-weight:700}
.foot{margin-top:10px;background:#222;color:#fff;text-align:center;padding:9px 16px}
.foot .s{font-family:'Bebas Neue';font-size:20px;color:#DCBE61}.foot .m{font-size:9px;opacity:.6;margin-top:4px}"""

HEAD = """<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Montserrat:wght@300;400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>%s</style></head><body><div class="page"><div class="bar"></div>"""

def rows_causa(causas, total):
    icon = lambda c: '🔧 ' if c in VEH_CAUSAS else ('👤 ' if 'CONDUCTOR' in c else '')
    out=""
    for c,n in causas.most_common():
        out+=f"<tr><td>{icon(c)}{c.title()}</td><td class='n'>{n}</td><td class='n'>{100*n/total:.0f}%</td></tr>"
    return out

def html_bus(d):
    he = HEAD % CSS
    he = he.replace('.hero::before{', f'.hero::before{{content:"{d["bus"]}";')
    causas_html = rows_causa(d['causas'], d['total_perd']) if d['total_perd'] else "<tr><td>Sin novedades</td><td class='n'>0</td><td></td></tr>"
    aus_html = "".join(f"<tr><td>{k.title()}</td><td class='n'>{v}</td></tr>" for k,v in d['aus_det'].most_common())
    mes_html = "".join(f"<tr><td>{MESES.get(m,'Sin fecha') if m!='??' else 'Sin fecha'}</td><td class='n'>{d['por_mes'][m]}</td></tr>" for m in sorted(d['por_mes']))
    dia_html = "".join(f"<tr><td>{DIAS[k]}</td><td class='n'>{v}</td></tr>" for k,v in sorted(d['por_dia'].items()))
    vuelta_rows = []
    for i, (k, v) in enumerate(sorted(d['por_vuelta'].items(), key=lambda x: (-x[1], x[0]))):
        style = 'style="background:#FFF0F0"' if i == 0 else ''
        warn = '⚠️ ' if i == 0 else ''
        label = 'Sin dato' if k == 0 else f'Vuelta {k}'
        vuelta_rows.append(f"<tr{style}><td>{warn}{label}</td><td class='n'>{v}</td></tr>")
    vuelta_html = "".join(vuelta_rows)
    ruta_rows = []
    for i, (r, v) in enumerate(d['por_ruta'].most_common(5)):
        style = 'style="background:#FFF0F0"' if i == 0 else ''
        warn = '⚠️ ' if i == 0 else ''
        ruta_rows.append(f"<tr{style}><td style='font-size:10px'>{warn}{r.title()}</td><td class='n'>{v}</td></tr>")
    ruta_html = "".join(ruta_rows)
    # Conductores por mes (de liquidaciones)
    cond_mes_html = "".join(
        f"<tr><td>{MESES.get(m[-2:],m)}</td><td class='n'>{len(conds)}</td><td style='font-size:10px'>{', '.join(c.title() for c in sorted(conds))}</td></tr>"
        for m, conds in sorted(d['cond_mes'].items())
    ) if d.get('cond_mes') else ""
    pveh = 100*d['veh_perd']/d['total_perd'] if d['total_perd'] else 0
    paus = 100*d['aus_perd']/d['total_perd'] if d['total_perd'] else 0
    tim_nota = f"L-V: {d['tim_sem']:.0f} · S-D: {d['tim_fds']:.0f} tim/vuelta · dato real del bus"
    return he + f"""
<div class="hero"><div class="hero-row"><div>
<span class="eyebrow">ANÁLISIS INDIVIDUAL DE BUSETA · 2026</span>
<h1 class="bebas">BUSETA <span class="r">{d['bus']}</span></h1>
<div class="slogan">"Transporte con corazón"</div></div>
<div class="meta">METROPOLITANA DE TRANSPORTES<br>LA CAROLINA · BARRANQUILLA<br>AFILIADO/PROP: {d['prop'].title()}</div>
</div></div>
<div class="body">
<div class="daterange">📅 <b>Período:</b> 1 ene – 31 may 2026. Registros de la buseta {d['bus']} del <b>{d['rango'][0]}</b> al <b>{d['rango'][1]}</b>.</div>
<h2>FICHA DEL VEHÍCULO</h2>
<div class="ficha">
<div class="f"><b>AFILIADO/PROP</b><span class="v">{d['prop'].title()}</span></div>
<div class="f"><b>PLACA</b><span class="v">{d['placa']}</span></div>
<div class="f"><b>MODELO</b><span class="v">{d['modelo']} · {d['edad']} años</span></div>
<div class="f"><b>RUTA</b><span class="v">{d['ruta'].title()}</span></div>
<div class="f"><b>DÍAS OPERADOS</b><span class="v">{d['ndias_op']} de 151</span></div>
<div class="f"><b>DÍAS SIN OPERAR</b><span class="v">{d['ndias_sin']}</span></div>
<div class="f" style="border-left:4px solid #DCBE61"><b>CONDUCTOR FIJO</b><span class="v">{d['conductor_fijo'].title()}</span></div>
<div class="f" style="border-left:4px solid #DCBE61"><b>DÍAS CON CONDUCTOR FIJO</b><span class="v">{d['conductor_fijo_dias']} días</span></div>
</div>
<div class="kpis">
<div class="k r"><div class="v">{d['total_perd']}</div><div class="l">vueltas perdidas en el período</div></div>
<div class="k r"><div class="v">{pveh:.0f}%</div><div class="l">de las pérdidas por el VEHÍCULO</div></div>
<div class="k"><div class="v">{d['taller']}</div><div class="l">días en taller (de 151)</div></div>
<div class="k"><div class="v">{d['conductores']}</div><div class="l">conductores distintos</div></div>
</div>
<h2>VUELTAS PERDIDAS POR CAUSA</h2>
<div class="row"><div>
<table><tr><th>Causa</th><th>Vueltas</th><th>%</th></tr>{causas_html}</table></div>
<div class="box d"><h3>AUSENCIA DEL CONDUCTOR · {d['aus_perd']} VUELTAS</h3>
<p style="font-size:10px;color:#555">Del {d['aus_rango'][0]} al {d['aus_rango'][1]} · motivo:</p>
<table>{aus_html or "<tr><td>Sin ausencias</td><td class='n'>0</td></tr>"}</table></div></div>
<h2>VUELTAS PERDIDAS — ¿CUÁNDO Y DÓNDE?</h2>
<div style="display:grid;grid-template-columns:1fr 1fr 1fr 1fr 1fr;gap:6px;margin-top:2px">
<div class="box r"><h3>MES</h3><p style="font-size:9px;color:#888;margin-bottom:3px">Vueltas perdidas</p><table>{mes_html}</table></div>
<div class="box r"><h3>DÍA</h3><p style="font-size:9px;color:#888;margin-bottom:3px">Vueltas perdidas</p><table>{dia_html}</table></div>
<div class="box r"><h3>VUELTA DEL DÍA</h3><p style="font-size:9px;color:#888;margin-bottom:3px">Vueltas perdidas (1=primera)</p><table>{vuelta_html or "<tr><td>Sin datos</td></tr>"}</table></div>
<div class="box r"><h3>RUTA</h3><p style="font-size:9px;color:#888;margin-bottom:3px">Vueltas perdidas</p><table>{ruta_html or "<tr><td>Sin datos</td></tr>"}</table></div>
<div class="box d"><h3>CONDUCTORES/MES</h3>
<table><tr><th>Mes</th><th>#</th><th>Conductor</th></tr>{cond_mes_html or "<tr><td colspan='3'>Sin datos</td></tr>"}</table></div>
</div>
<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:6px">
<div class="diag"><h3>🧠 RESUMEN OPERATIVO</h3><ul>
<li>De {d['total_perd']} vueltas perdidas, <b>{pveh:.0f}% son del vehículo</b> (taller+varado) y {paus:.0f}% por ausencia del conductor.</li>
<li>Bus modelo {d['modelo']} · {d['taller']} días en taller · operó {d['ndias_op']} de 151 días.</li>
<li>Ausencias del conductor: {", ".join(f"{k.lower()} ({v})" for k,v in d['aus_det'].most_common(3)) or "sin ausencias registradas"}.</li>
</ul></div>
<div class="diag" style="background:#FFF0F0;border-left:5px solid #C22219"><h3 style="color:#C22219">📉 INGRESO BRUTO NO GENERADO — FALLAS DEL VEHÍCULO</h3><ul>
<li>Fallas del vehículo (taller + varado): <b>{d['veh_perd']} vueltas</b> · promedio flota: {FLOTA_VEH_PROM} vueltas/bus {'⚠️' if d['veh_perd'] > FLOTA_VEH_PROM else '✅'}.</li>
<li>Timbradas no registradas: <b>{d['timbradas_perdidas']:,} timbradas</b> ({d['veh_perd']} vueltas × {tim_nota}).</li>
<li>Ingreso bruto no generado por el bus: <b>${d['ingreso_perdido']:,.0f} COP</b> ({d['timbradas_perdidas']:,} × $3.300).</li>
<li>Mantenimiento oportuno = más vueltas = más timbradas para todos.</li>
</ul></div>
</div>
</div>
<div class="foot"><div class="s bebas">"TRANSPORTE CON CORAZÓN"</div>
<div class="m">Metropolitana de Transportes La Carolina · Reporte buseta {d['bus']} · Fuentes: Flota, Liquidaciones, Viajes perdidos</div></div>
</div></body></html>"""

def html_conductor(d):
    he = HEAD % CSS
    he = he.replace('.hero::before{', '.hero::before{content:"C";')
    causas_html = rows_causa(d['causas'], d['total_perd']) if d['total_perd'] else "<tr><td>Sin novedades</td><td class='n'>0</td><td></td></tr>"
    aus_html = "".join(f"<tr><td>{k.title()}</td><td class='n'>{v}</td></tr>" for k,v in d['aus_det'].most_common())
    buses_html = "".join(f"<tr><td>Bus {b}</td><td class='n'>{n}</td></tr>" for b,n in d['buses'].most_common(8))
    rutas_html = "".join(f"<tr><td>{r.title()}</td><td class='n'>{n}</td></tr>" for r,n in d['rutas'].most_common(5))
    nom = d['nombre'].title()
    return he + f"""
<div class="hero"><div class="hero-row"><div>
<span class="eyebrow">PERFIL DE CONDUCTOR · 2026</span>
<h1 class="bebas" style="font-size:26px">{nom}</h1>
<div class="slogan">"Transporte con corazón"</div></div>
<div class="meta">METROPOLITANA DE TRANSPORTES<br>LA CAROLINA · BARRANQUILLA<br>CÉDULA: {d['ced']}</div>
</div></div>
<div class="body">
<div class="daterange">📅 <b>Período:</b> 1 ene – 31 may 2026. Registros de {nom} del <b>{d['rango'][0]}</b> al <b>{d['rango'][1]}</b>.</div>
<div class="kpis">
<div class="k"><div class="v">{d['ndias']}</div><div class="l">días con operación registrada</div></div>
<div class="k"><div class="v">{len(d['buses'])}</div><div class="l">buses distintos manejados</div></div>
<div class="k r"><div class="v">{d['total_perd']}</div><div class="l">vueltas perdidas asociadas</div></div>
<div class="k r"><div class="v">{d['causas'].get('AUSENCIA CONDUCTOR',0)}</div><div class="l">por ausencia del conductor</div></div>
</div>
<h2>BUSES QUE MANEJÓ</h2>
<div class="row">
<div class="box d"><h3>BUSES</h3><table>{buses_html}</table></div>
<div class="box d"><h3>RUTAS</h3><table>{rutas_html}</table></div></div>
<h2>NOVEDADES / VUELTAS PERDIDAS</h2>
<div class="row"><div>
<table><tr><th>Causa</th><th>Vueltas</th><th>%</th></tr>{causas_html}</table></div>
<div class="box d"><h3>SI HUBO AUSENCIA · MOTIVO</h3>
<table>{aus_html or "<tr><td>Sin ausencias</td><td class='n'>0</td></tr>"}</table></div></div>
<div class="diag"><h3>🧠 RESUMEN</h3><ul>
<li>{nom} (CC {d['ced']}) operó <b>{d['ndias']} días</b> en <b>{len(d['buses'])} buses</b> distintos.</li>
<li>Vueltas perdidas asociadas: {d['total_perd']} · ausencias del conductor: {d['causas'].get('AUSENCIA CONDUCTOR',0)}.</li>
<li>Motivos de ausencia: {", ".join(f"{k.lower()} ({v})" for k,v in d['aus_det'].most_common(3)) or "—"}.</li>
</ul></div>
</div>
<div class="foot"><div class="s bebas">"TRANSPORTE CON CORAZÓN"</div>
<div class="m">Metropolitana de Transportes La Carolina · Perfil de conductor · Fuentes: Liquidaciones, Viajes perdidos</div></div>
</div></body></html>"""

def render(html, nombre):
    os.makedirs(OUT_DIR, exist_ok=True)
    htmlpath = os.path.join(OUT_DIR, nombre+".html")
    pdfpath = os.path.join(OUT_DIR, nombre+".pdf")
    open(htmlpath,"w", encoding="utf-8").write(html)
    result = subprocess.run([CHROME,"--headless","--disable-gpu","--no-sandbox","--no-pdf-header-footer",
        "--virtual-time-budget=9000",f"--print-to-pdf={pdfpath}",f"file://{htmlpath}"],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=60)
    if result.returncode != 0 or not os.path.exists(pdfpath):
        raise RuntimeError(f"No se pudo generar PDF con Chromium: {result.stderr.strip() or result.stdout.strip()}")
    return pdfpath

def main():
    if len(sys.argv) < 3:
        print("Uso: python3 generar_reporte.py [bus|conductor] <valor>"); sys.exit(1)
    require_runtime_inputs()
    modo, valor = sys.argv[1].lower(), " ".join(sys.argv[2:])
    if modo == "bus":
        d = analizar_bus(valor.strip())
        if d['total_perd']==0 and d['ndias_op']==0:
            print(f"⚠️ No encontré datos del bus {valor}"); sys.exit(0)
        pdf = render(html_bus(d), f"REPORTE_BUSETA_{valor.strip()}")
        print("RESUMEN BUS", valor, ":", d['prop'], "| modelo", d['modelo'],
              "|", d['total_perd'], "vueltas perdidas |", d['taller'], "días taller |", d['conductores'], "conductores")
        print("PDF:", pdf)
    elif modo == "conductor":
        d = analizar_conductor(valor.strip())
        if not d: print(f"⚠️ No encontré al conductor '{valor}'"); sys.exit(0)
        safe = re.sub(r'[^A-Za-z0-9]+','_', d['nombre'])[:40]
        pdf = render(html_conductor(d), f"REPORTE_CONDUCTOR_{safe}")
        print("RESUMEN CONDUCTOR:", d['nombre'], "CC", d['ced'], "|", d['ndias'], "días |",
              len(d['buses']), "buses |", d['total_perd'], "vueltas perdidas")
        print("PDF:", pdf)
    else:
        print("Modo inválido. Usa 'bus' o 'conductor'.")

if __name__ == "__main__":
    main()
